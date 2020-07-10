"""官方文档：https://openpyxl.readthedocs.io/en/stable/index.html"""
import openpyxl

class ReadExcel:
    """从excel中读取测试数据"""
    def read_excel(self):

        wb = openpyxl.load_workbook("test_data.xlsx")
        ws = wb["test_data"]

        test_data = []  # 存储测试数据

        for i in range(2, ws.max_row+1):
            # 存储每行的测试数据
            sub_data = {}
            sub_data["case_id"] = ws.cell(i, 1).value
            sub_data["title"] = ws.cell(i, 2).value
            sub_data["method"] = ws.cell(i, 3).value
            sub_data["url"] = ws.cell(i, 4).value
            sub_data["params"] = ws.cell(i, 5).value
            sub_data["expected_code"] = ws.cell(i, 6).value
            sub_data["response"] = ws.cell(i, 7).value
            sub_data["result"] = ws.cell(i, 8).value

            test_data.append(sub_data)

        return test_data