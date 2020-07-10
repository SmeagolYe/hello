"""官方文档：https://openpyxl.readthedocs.io/en/stable/index.html"""
import openpyxl
from common.read_conf import ReadConfig
from common import pro_path

class ReadExcel:
    def __init__(self, file_name):
        self.file_name = file_name
        # self.sheet_name = sheet_name

    """从excel中读取测试数据"""
    def read_excel(self):

        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["test_data"]

        """从excel里面获取初始化手机号"""
        tel = self.get_tel()

        """记录使用过的号码"""
        self.used_tel(tel)

        """获取初始化手机号后，更新初始化的手机号码"""
        self.write_back(1, 1, tel + 1, "tel")

        """用来存储excel中所有的用例"""
        test_data = []

        for i in range(2, ws.max_row+1):
            sub_data = {}
            sub_data["case_id"] = ws.cell(i, 1).value
            sub_data["title"] = ws.cell(i, 2).value
            sub_data["method"] = ws.cell(i, 3).value
            sub_data["url"] = ws.cell(i, 4).value

            """如果请求参数中带有${tel}，则替换为从excel读取的手机号"""
            if ws.cell(i, 5).value.find("${tel}") != -1:
                sub_data["params"] = ws.cell(i, 5).value.replace("${tel}", str(tel))
            else:
                sub_data["params"] = ws.cell(i, 5).value

            print(sub_data["params"])
            sub_data["expected_code"] = ws.cell(i, 6).value
            sub_data["response"] = ws.cell(i, 7).value
            sub_data["result"] = ws.cell(i, 8).value

            test_data.append(sub_data)

        """用来存储最终需要被执行的用例"""
        final_data = []
        mode = ReadConfig().read_config(pro_path.conf_path, "MODE", "mode")
        case_ids = eval(ReadConfig().read_config(pro_path.conf_path, "CASE_IDS", "case_ids"))

        """若mode值为on，则执行所有用例，否则只执行指定的用例"""
        if mode == "on":
            final_data = test_data
        else:
            for i in case_ids:
                final_data.append(test_data[i - 1])

        return final_data

    """用例结果回写"""
    def write_back(self, row, col, new_value, sheet_name):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb[sheet_name]

        ws.cell(row, col).value = new_value
        wb.save(self.file_name)

    """获取Excel的sheet名为tel的初始化手机号"""
    def get_tel(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["tel"]
        tel = ws.cell(1, 1).value
        return tel

    def used_tel(self, tel):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb["used_tel"]
        ws.cell(ws.max_row + 1, 1).value = tel
        wb.save(self.file_name)

    # """更新初始化手机号，每读取一次手机号，手机号+1"""
    # def update_tel(self, new_tel):
    #     wb = openpyxl.load_workbook(self.file_name)
    #     ws = wb["tel"]
    #     ws.cell(1, 1).value = new_tel
    #     wb.save(self.file_name)

